import xlrd
import openpyxl
import statistics

filename = 'stud_detail.xlsx'
wb = xlrd.open_workbook(filename, encoding_override="utf-8") 
detailfile = wb.sheet_by_index(0)

Checker = {}
response = []
entries = len(detailfile.col(0))
i = 1
while i < entries:
    indexque = 2
    indexans = 3
    while indexque <= 30:
        q_no = int(detailfile.cell_value(i,indexque))
        ans = int(detailfile.cell_value(i,indexans))
        indexque += 3
        indexans += 3

        if(q_no not in Checker.keys()):
            Checker[q_no] = []
            Checker[q_no].append(ans)
        else:
            Checker[q_no].append(ans)

    i += 1


def find_max_mode(list1):
    list_table = statistics._counts(list1)
    len_table = len(list_table)

    if len_table == 1:
        max_mode = statistics.mode(list1)
    else:
        new_list = []
        for i in range(len_table):
            new_list.append(list_table[i][0])
        max_mode = max(new_list) # use the max value here
    return max_mode





for que,ans in Checker.items():
    mode = find_max_mode(ans)
    # print("mode of question {} is {}".format(que,mode))
    Checker[que] = mode



filename = 'questions.xlsx'
wb = openpyxl.load_workbook(filename)
sheet = wb['questions']
c_val = (len(sheet['G']))

#update level
# for z in range(1, c_val+1):
#     if z in Checker.keys():
#         if sheet.cell(row = z, column = 7 ).value == Checker.values[z]:
#             print('will remain same')
#         elif sheet.cell(row = z, column = 7 ).value == 1 and Checker.values[z] == 0:
            

   
