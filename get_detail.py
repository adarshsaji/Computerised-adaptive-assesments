import openpyxl      #to write from the excel sheet

filename = 'stud_detail.xlsx'
wb = openpyxl.load_workbook(filename)
sheet = wb['Sheet1']
c_val = (len(sheet['A']))

def get_data():
    name = input("Enter your name: ")
    roll_number = (input("Enter your roll number: "))
   

    sheet.cell(row=c_val+1 , column=1 ).value = name
    sheet.cell(row=c_val+1 , column=2 ).value = roll_number

    wb.save(filename)

def put_data(responses):
    studentResponse, timeTaken = responses
    i = 3       #for data entry from 3rd column
    j = 0       #for time taken index
    for key,value in studentResponse.items():
         sheet.cell(row = c_val+1 , column = i ).value = key
         i += 1 
         sheet.cell(row = c_val+1 , column = i ).value = value
         i += 1
         sheet.cell(row = c_val+1 , column = i ).value = timeTaken[j]
         i += 1
         j += 1
                 
    wb.save(filename)